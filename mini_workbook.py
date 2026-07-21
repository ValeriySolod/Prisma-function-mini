"""Atomic publication of the authoritative Prisma Function Mini workbook."""

from __future__ import annotations

import os
import tempfile
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from mini_domain import MiniOutputRow, OUTPUT_COLUMNS, WORKSHEET_NAME
from mini_storage import MiniAuctionStorage


class MiniWorkbookError(RuntimeError):
    """The cumulative workbook could not be published safely."""


class MiniWorkbookPublisher:
    """Render a deterministic workbook snapshot from Mini's SQLite history."""

    COLUMN_WIDTHS = {
        "Auction Date": 15,
        "Exit Market / Storage": 24,
        "Entry Market / Storage": 24,
        "Capacity Type": 15,
        "Network Point": 36,
        "Product Type": 14,
        "Flow Start": 21,
        "Flow End": 21,
        "Booked Capacity (kWh/h)": 27,
        "Duration (hours)": 18,
        "Auction Tariff (EUR/MWh/h)": 29,
    }
    NUMBER_FORMATS = {
        1: "yyyy-mm-dd",
        7: "yyyy-mm-dd hh:mm",
        8: "yyyy-mm-dd hh:mm",
        9: "0.############################",
        10: "0.############################",
        11: "0.############################",
    }
    WIDTH_TOLERANCE = 1e-6

    def __init__(self, storage: MiniAuctionStorage) -> None:
        if not isinstance(storage, MiniAuctionStorage):
            raise TypeError("storage must be MiniAuctionStorage.")
        approved_result = (
            storage.paths.root / "data" / "result" / "prisma_function_mini.xlsx"
        )
        if storage.paths.result != approved_result:
            raise ValueError("Mini workbook must use the approved runtime path.")
        self.storage = storage
        self.output_path = storage.paths.result

    def publish(self) -> Path:
        """Atomically replace the workbook, or leave an equivalent one untouched."""

        history = self.storage.history()
        expected = tuple(MiniOutputRow.from_record(item.auction).values() for item in history)
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        staged: Path | None = None
        try:
            descriptor, name = tempfile.mkstemp(
                prefix=f".{self.output_path.stem}-", suffix=".xlsx",
                dir=self.output_path.parent,
            )
            os.close(descriptor)
            staged = Path(name)
            self._write(staged, expected)
            if not self.validate(staged, expected):
                raise MiniWorkbookError("The staged Excel workbook failed validation.")
            if self.output_path.exists() and self.validate(self.output_path, expected):
                return self.output_path
            try:
                os.replace(staged, self.output_path)
            except PermissionError as exc:
                raise MiniWorkbookError(
                    "The Excel output is open or locked. Close it and retry publication."
                ) from exc
            staged = None
            return self.output_path
        except MiniWorkbookError:
            raise
        except Exception as exc:
            raise MiniWorkbookError("The Excel output could not be staged safely.") from exc
        finally:
            if staged is not None:
                try:
                    staged.unlink(missing_ok=True)
                except OSError:
                    pass

    @classmethod
    def _write(cls, path: Path, rows: Iterable[tuple[object, ...]]) -> None:
        workbook = Workbook()
        try:
            sheet = workbook.active
            sheet.title = WORKSHEET_NAME
            sheet.append(OUTPUT_COLUMNS)
            for row in rows:
                sheet.append(tuple(
                    float(value) if isinstance(value, Decimal) else value
                    for value in row
                ))
            for index, header in enumerate(OUTPUT_COLUMNS, start=1):
                sheet.column_dimensions[get_column_letter(index)].width = cls.COLUMN_WIDTHS[header]
            for column, number_format in cls.NUMBER_FORMATS.items():
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row=row, column=column).number_format = number_format
            workbook.save(path)
        finally:
            workbook.close()

    @classmethod
    def validate(cls, path: Path, expected_rows: Iterable[tuple[object, ...]]) -> bool:
        workbook = None
        try:
            workbook = load_workbook(path, read_only=False, data_only=True)
            if workbook.sheetnames != [WORKSHEET_NAME]:
                return False
            sheet = workbook[WORKSHEET_NAME]
            headers = tuple(cell.value for cell in sheet[1])
            if headers != OUTPUT_COLUMNS:
                return False
            if any(
                abs(
                    sheet.column_dimensions[get_column_letter(index)].width
                    - cls.COLUMN_WIDTHS[header]
                )
                > cls.WIDTH_TOLERANCE
                for index, header in enumerate(OUTPUT_COLUMNS, start=1)
            ):
                return False
            actual = tuple(tuple(cell.value for cell in row) for row in sheet.iter_rows(min_row=2))
            expected = tuple(cls._excel_values(row) for row in expected_rows)
            if actual != expected:
                return False
            return all(
                sheet.cell(row=row, column=column).number_format == number_format
                for column, number_format in cls.NUMBER_FORMATS.items()
                for row in range(2, sheet.max_row + 1)
            )
        except Exception:
            return False
        finally:
            if workbook is not None:
                workbook.close()

    @staticmethod
    def _excel_values(row: tuple[object, ...]) -> tuple[object, ...]:
        return tuple(
            float(value) if isinstance(value, Decimal)
            else datetime.combine(value, datetime.min.time()) if type(value) is date
            else value
            for value in row
        )
