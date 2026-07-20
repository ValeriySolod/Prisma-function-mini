from __future__ import annotations

from datetime import date, datetime, timezone
from pathlib import Path

import hashlib
import json
import pandas as pd
import pytest
import sqlite3
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import prisma_import_workflow as workflow
from csv_contracts import MONITORING_CSV_COLUMNS, PRISMA_EXPORT_COLUMNS, CsvDetectionResult, CsvFormat
from prisma_import_workflow import PrismaWorkflowError, run_prisma_import_workflow
from processor import import_prisma_export
from storage import AuctionStorage, AuctionStorageError


BASE = {
    "Auction ID": "A-1", "Start of Auction": "01.01.2025 09:00",
    "Marketed Capacity": "1000", "Unit Marketed Capacity": "kWh/h",
    "Product Runtime Start": "02.01.2025 00:00", "Product Runtime End": "03.01.2025 00:00",
    "Direction": "Entry", "Network Point Name Entry": "VGS Storage Hub (4290)",
    "Network Point ID Entry": "ENTRY-ID", "Regulated Tariff Entry TSO": "0.75",
    "Unit Regulated Entry Capacity Tariff": "cent/kWh/h/Runtime",
    "Surcharge": "0.5", "Unit Surcharge": "cent/kWh/h/Runtime", "State": "Open",
}


def write_export(path: Path, rows: list[dict]) -> Path:
    pd.DataFrame(rows).reindex(columns=PRISMA_EXPORT_COLUMNS).fillna("").to_csv(
        path, sep=";", encoding="cp1252", index=False
    )
    return path


def run(source: Path, root: Path, day: date):
    return run_prisma_import_workflow(
        source, source_date=day, evaluated_at=datetime(2025, 1, 10, tzinfo=timezone.utc),
        database_path=root / "auctions.db", state_path=root / "state.json",
        output_path=root / "result.xlsx",
    )


def valid_workbook_bytes(path: Path) -> bytes:
    pd.DataFrame(columns=AuctionStorage.EXCEL_COLUMNS).to_excel(
        path, index=False, sheet_name="Auctions"
    )
    AuctionStorage.apply_excel_widths(path)
    assert AuctionStorage.validate_excel(path)
    return path.read_bytes()


def test_new_repeat_and_next_daily_export_are_cumulative_and_enriched(tmp_path):
    first = write_export(tmp_path / "first.csv", [BASE])
    initial = run(first, tmp_path, date(2025, 1, 1))
    assert (initial.processed, initial.inserted, initial.updated, initial.unchanged) == (1, 1, 0, 0)
    repeated = run(first, tmp_path, date(2025, 1, 1))
    assert repeated.source_status.value == "unchanged"
    # Exact retry reports the persisted historical import summary.
    assert (repeated.inserted, repeated.updated) == (1, 0)

    next_rows = [
        {**BASE, "State": "Finished"},
        {**BASE, "Auction ID": "A-2", "Network Point ID Entry": "ENTRY-2"},
        {**BASE, "Auction ID": "A-3", "Network Point ID Entry": "ENTRY-3"},
    ]
    second = write_export(tmp_path / "second.csv", next_rows)
    daily = run(second, tmp_path, date(2025, 1, 2))
    assert (daily.processed, daily.inserted, daily.updated, daily.unchanged) == (3, 2, 1, 0)
    frame = pd.read_excel(tmp_path / "result.xlsx")
    assert frame["Entry Market/Storage"].tolist() == ["VGS Storage Hub"] * 3
    assert frame["Auction ID"].tolist() == ["A-1", "A-2", "A-3"]


def test_normal_import_never_runs_historical_backfill(tmp_path, monkeypatch):
    monkeypatch.setattr(
        AuctionStorage,
        "backfill_historical_market_storage",
        lambda *_: pytest.fail("historical backfill must remain explicit"),
    )
    result = run(
        write_export(tmp_path / "source.csv", [BASE]),
        tmp_path,
        date(2025, 1, 1),
    )
    assert result.inserted == 1


def test_daily_export_reports_inserted_updated_and_unchanged(tmp_path):
    first_rows = [BASE, {**BASE, "Auction ID": "A-2", "Network Point ID Entry": "E-2"}]
    run(write_export(tmp_path / "one.csv", first_rows), tmp_path, date(2025, 1, 1))
    next_rows = [BASE, {**first_rows[1], "State": "Finished"}, {**BASE, "Auction ID": "A-3", "Network Point ID Entry": "E-3"}]
    result = run(write_export(tmp_path / "two.csv", next_rows), tmp_path, date(2025, 1, 2))
    assert (result.inserted, result.updated, result.unchanged) == (1, 1, 1)


def test_malformed_rows_are_audited_without_losing_valid_rows(tmp_path):
    source = write_export(tmp_path / "mixed.csv", [BASE, {**BASE, "Direction": "sideways"}])
    result = run(source, tmp_path, date(2025, 1, 1))
    assert (result.processed, result.rejected, len(result.issues)) == (1, 1, 1)
    assert result.issues[0].source_row_number == 3


def test_monitoring_and_unsupported_inputs_are_rejected(tmp_path):
    monitoring = tmp_path / "monitor.csv"
    monitoring.write_text(",".join(MONITORING_CSV_COLUMNS) + "\n", encoding="utf-8")
    with pytest.raises(PrismaWorkflowError) as caught:
        run(monitoring, tmp_path, date(2025, 1, 1))
    assert str(caught.value) == (
        "Monitoring CSV cannot be imported as detailed PRISMA results. "
        "Use Load Monitoring CSV for live monitoring."
    )
    unknown = tmp_path / "unknown.csv"
    unknown.write_text("name,value\n", encoding="utf-8")
    with pytest.raises(PrismaWorkflowError, match="Unsupported CSV format"):
        run(unknown, tmp_path, date(2025, 1, 1))


def test_ambiguous_input_is_rejected_explicitly(tmp_path, monkeypatch):
    source = tmp_path / "ambiguous.csv"
    source.write_text("anything\n", encoding="utf-8")
    monkeypatch.setattr(workflow, "detect_csv_format", lambda path: CsvDetectionResult(CsvFormat.AMBIGUOUS))
    with pytest.raises(PrismaWorkflowError, match="ambiguous"):
        run(source, tmp_path, date(2025, 1, 1))


def test_output_is_deterministic_for_reversed_input(tmp_path):
    rows = [BASE, {**BASE, "Auction ID": "A-2", "Network Point ID Entry": "E-2"}]
    first_root, second_root = tmp_path / "a", tmp_path / "b"
    first_root.mkdir(); second_root.mkdir()
    run(write_export(first_root / "x.csv", rows), first_root, date(2025, 1, 1))
    run(write_export(second_root / "x.csv", list(reversed(rows))), second_root, date(2025, 1, 1))
    left = pd.read_excel(first_root / "result.xlsx").fillna("").to_dict("records")
    right = pd.read_excel(second_root / "result.xlsx").fillna("").to_dict("records")
    assert left == right


def test_excel_publication_failure_preserves_previous_bytes_and_retry_recovers(tmp_path, monkeypatch):
    source = write_export(tmp_path / "source.csv", [BASE])
    output = tmp_path / "result.xlsx"
    previous = valid_workbook_bytes(output)
    real_replace = workflow.AuctionStorage.export_excel.__globals__["os"].replace
    monkeypatch.setattr(
        workflow.AuctionStorage.export_excel.__globals__["os"], "replace",
        lambda *_: (_ for _ in ()).throw(PermissionError("locked")),
    )
    with pytest.raises(PrismaWorkflowError, match="open or locked"):
        run(source, tmp_path, date(2025, 1, 1))
    assert output.read_bytes() == previous
    with sqlite3.connect(tmp_path / "auctions.db") as connection:
        assert connection.execute("SELECT count(*) FROM auctions").fetchone()[0] == 1
        assert connection.execute("SELECT status FROM prisma_source_operations").fetchone()[0] == "data_committed"
    assert not list(tmp_path.glob(".result-*.xlsx"))

    monkeypatch.setattr(workflow.AuctionStorage.export_excel.__globals__["os"], "replace", real_replace)
    retried = run(source, tmp_path, date(2025, 1, 1))
    assert retried.inserted == 1
    assert workflow.AuctionStorage.validate_excel(output)


def test_different_same_date_source_is_blocked_while_operation_unresolved(tmp_path, monkeypatch):
    first = write_export(tmp_path / "first.csv", [BASE])
    monkeypatch.setattr(workflow.AuctionStorage, "export_excel", lambda *_: (_ for _ in ()).throw(workflow.AuctionStorageError("stage failed")))
    with pytest.raises(PrismaWorkflowError, match="stage failed"):
        run(first, tmp_path, date(2025, 1, 1))
    changed = write_export(tmp_path / "changed.csv", [{**BASE, "State": "Finished"}])
    with pytest.raises(PrismaWorkflowError, match="different PRISMA source.*unresolved"):
        run(changed, tmp_path, date(2025, 1, 1))


@pytest.mark.parametrize("damage", ["missing", "corrupt"])
def test_exact_retry_repairs_missing_or_corrupt_output_without_mutating_rows(tmp_path, damage):
    source = write_export(tmp_path / "source.csv", [BASE])
    initial = run(source, tmp_path, date(2025, 1, 1))
    output = tmp_path / "result.xlsx"
    if damage == "missing":
        output.unlink()
    else:
        output.write_bytes(b"not an xlsx")
    retried = run(source, tmp_path, date(2025, 1, 1))
    assert retried.source_status is workflow.SourceUpdateStatus.UNCHANGED
    assert retried.inserted == initial.inserted == 1
    assert workflow.AuctionStorage.validate_excel(output)
    with sqlite3.connect(tmp_path / "auctions.db") as connection:
        assert connection.execute("SELECT count(*) FROM auctions").fetchone()[0] == 1


def test_exact_retry_repairs_legacy_default_widths_without_mutating_rows(tmp_path):
    source = write_export(tmp_path / "source.csv", [BASE])
    initial = run(source, tmp_path, date(2025, 1, 1))
    initial_counts = (
        initial.processed, initial.inserted, initial.updated, initial.unchanged,
        initial.filtered, initial.rejected, initial.audit_issue_count,
    )
    output = tmp_path / "result.xlsx"
    workbook = load_workbook(output)
    sheet = workbook["Auctions"]
    for index in range(1, len(AuctionStorage.EXCEL_COLUMNS) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 13
    workbook.save(output)
    workbook.close()
    assert not AuctionStorage.validate_excel(output)

    storage = AuctionStorage(tmp_path / "auctions.db")
    operations_before = storage.operations()
    assert len(operations_before) == 1
    operation_before = dict(operations_before[0])
    with sqlite3.connect(storage.database_path) as connection:
        auctions_before = connection.execute(
            "SELECT * FROM auctions ORDER BY id"
        ).fetchall()
    retried = run(source, tmp_path, date(2025, 1, 1))
    operations_after = storage.operations()
    assert len(operations_after) == 1
    operation_after = dict(operations_after[0])
    with sqlite3.connect(storage.database_path) as connection:
        auctions_after = connection.execute(
            "SELECT * FROM auctions ORDER BY id"
        ).fetchall()

    assert retried.source_status is workflow.SourceUpdateStatus.UNCHANGED
    assert auctions_after == auctions_before
    assert operation_after == operation_before
    assert (
        retried.processed, retried.inserted, retried.updated, retried.unchanged,
        retried.filtered, retried.rejected, retried.audit_issue_count,
    ) == initial_counts
    assert AuctionStorage.validate_excel(output)


def test_header_only_export_is_accepted_as_distinct_empty_import(tmp_path):
    result = run(write_export(tmp_path / "empty.csv", []), tmp_path, date(2025, 1, 1))
    assert (result.processed, result.filtered, result.rejected, result.inserted) == (0, 0, 0, 0)
    assert result.source_status is workflow.SourceUpdateStatus.APPLIED


@pytest.mark.parametrize(
    ("row", "filtered", "rejected"),
    [({**BASE, "Marketed Capacity": "999"}, 1, 0),
     ({**BASE, "Direction": "sideways"}, 0, 1)],
)
def test_fully_nonimportable_exports_keep_exact_audit_counts(
    tmp_path, row, filtered, rejected
):
    result = run(write_export(tmp_path / "source.csv", [row]), tmp_path, date(2025, 1, 1))
    assert (result.processed, result.filtered, result.rejected) == (0, filtered, rejected)
    assert result.audit_issue_count == 1
    retried = run(tmp_path / "source.csv", tmp_path, date(2025, 1, 1))
    assert (retried.filtered, retried.rejected, retried.audit_issue_count) == (
        filtered, rejected, 1
    )


def test_sqlite_failure_before_pending_record_leaves_everything_untouched(tmp_path, monkeypatch):
    source = write_export(tmp_path / "source.csv", [BASE])
    monkeypatch.setattr(
        workflow.AuctionStorage, "begin_operation",
        lambda *_: (_ for _ in ()).throw(AuctionStorageError("begin failed")),
    )
    with pytest.raises(PrismaWorkflowError, match="begin failed"):
        run(source, tmp_path, date(2025, 1, 1))
    with sqlite3.connect(tmp_path / "auctions.db") as connection:
        assert connection.execute("SELECT count(*) FROM auctions").fetchone()[0] == 0
        assert connection.execute("SELECT count(*) FROM prisma_source_operations").fetchone()[0] == 0
    assert not (tmp_path / "result.xlsx").exists()


def test_sqlite_mid_transaction_failure_rolls_back_and_retry_resumes(tmp_path, monkeypatch):
    source = write_export(tmp_path / "source.csv", [BASE])
    original = AuctionStorage._upsert_rows

    def fail_after_mutation(connection, rows):
        original(connection, rows)
        raise sqlite3.OperationalError("injected transaction failure")

    monkeypatch.setattr(AuctionStorage, "_upsert_rows", staticmethod(fail_after_mutation))
    with pytest.raises(PrismaWorkflowError, match="injected"):
        run(source, tmp_path, date(2025, 1, 1))
    with sqlite3.connect(tmp_path / "auctions.db") as connection:
        assert connection.execute("SELECT count(*) FROM auctions").fetchone()[0] == 0
        assert connection.execute("SELECT status FROM prisma_source_operations").fetchone()[0] == "pending"
    monkeypatch.setattr(AuctionStorage, "_upsert_rows", staticmethod(original))
    assert run(source, tmp_path, date(2025, 1, 1)).inserted == 1


@pytest.mark.parametrize("boundary", ["staging", "validation"])
def test_excel_prepublication_failures_preserve_previous_output_and_retry(
    tmp_path, monkeypatch, boundary
):
    source = write_export(tmp_path / "source.csv", [BASE])
    output = tmp_path / "result.xlsx"
    previous = valid_workbook_bytes(output)
    if boundary == "staging":
        original = pd.DataFrame.to_excel
        monkeypatch.setattr(
            pd.DataFrame, "to_excel",
            lambda *_args, **_kwargs: (_ for _ in ()).throw(OSError("disk full")),
        )
    else:
        original = AuctionStorage.validate_excel
        monkeypatch.setattr(AuctionStorage, "validate_excel", staticmethod(lambda _: False))
    with pytest.raises(PrismaWorkflowError):
        run(source, tmp_path, date(2025, 1, 1))
    assert output.read_bytes() == previous
    assert not list(tmp_path.glob(".result-*.xlsx"))
    with sqlite3.connect(tmp_path / "auctions.db") as connection:
        assert connection.execute("SELECT status FROM prisma_source_operations").fetchone()[0] == "data_committed"
    if boundary == "staging":
        monkeypatch.setattr(pd.DataFrame, "to_excel", original)
    else:
        monkeypatch.setattr(AuctionStorage, "validate_excel", staticmethod(original))
    assert run(source, tmp_path, date(2025, 1, 1)).source_status is workflow.SourceUpdateStatus.APPLIED


def test_finalization_failure_remains_recoverable_and_never_reports_success(tmp_path, monkeypatch):
    source = write_export(tmp_path / "source.csv", [BASE])
    original = AuctionStorage.finalize_operation
    monkeypatch.setattr(
        AuctionStorage, "finalize_operation",
        lambda *_: (_ for _ in ()).throw(AuctionStorageError("finalize failed")),
    )
    with pytest.raises(PrismaWorkflowError, match="finalize failed"):
        run(source, tmp_path, date(2025, 1, 1))
    assert AuctionStorage.validate_excel(tmp_path / "result.xlsx")
    with sqlite3.connect(tmp_path / "auctions.db") as connection:
        assert connection.execute("SELECT status FROM prisma_source_operations").fetchone()[0] == "data_committed"
    monkeypatch.setattr(AuctionStorage, "finalize_operation", original)
    recovered = run(source, tmp_path, date(2025, 1, 1))
    assert recovered.inserted == 1
    with sqlite3.connect(tmp_path / "auctions.db") as connection:
        assert connection.execute("SELECT status FROM prisma_source_operations").fetchone()[0] == "accepted"


def test_legacy_json_migrates_with_unavailable_metadata_and_repairs_output(tmp_path):
    source = write_export(tmp_path / "source.csv", [BASE])
    digest = hashlib.sha256(source.read_bytes()).hexdigest()
    AuctionStorage(tmp_path / "auctions.db").upsert(import_prisma_export(source).rows)
    (tmp_path / "state.json").write_text(json.dumps({"accepted_sources": [{
        "source_date": "2025-01-01", "source_name": source.name, "sha256": digest,
    }]}), encoding="utf-8")
    result = run(source, tmp_path, date(2025, 1, 1))
    assert result.source_status is workflow.SourceUpdateStatus.UNCHANGED
    assert result.processed is result.filtered is result.rejected is None
    assert "unavailable" in result.summary()
    assert AuctionStorage.validate_excel(tmp_path / "result.xlsx")
