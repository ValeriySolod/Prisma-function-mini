import sqlite3
import threading
from contextlib import closing

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from storage import AuctionStorage, AuctionStorageError, HistoricalBackfillStatus


APPROVED_EXCEL_COLUMN_WIDTHS = {
    "Auction Date": 21,
    "Exit Market/Storage": 22,
    "Entry Market/Storage": 22,
    "Capacity Type": 15,
    "Network Point Name": 36,
    "Product Type": 14,
    "Flow Start": 21,
    "Flow End": 21,
    "Booked Capacity, kWh/h": 24,
    "Runtime Hours": 15,
    "Tariff, EUR/MWh/h": 20,
    "Premium, EUR/MWh/h": 21,
    "Auction ID": 16,
    "TSO Exit": 30,
    "TSO Entry": 30,
    "Status": 14,
}


def test_duplicate_import_does_not_insert_twice(tmp_path) -> None:
    storage = AuctionStorage(tmp_path / "test.db")
    row = {
        "auction_id": "123",
        "auction_date": "2026-07-10T06:00:00",
        "exit_market": "",
        "entry_market": "",
        "direction": "entry",
        "network_point": "Test Point",
        "network_point_id": "NP-1",
        "tso_exit": "",
        "tso_entry": "Test TSO",
        "product_type": "Day Ahead",
        "flow_start": "2026-07-10T06:00:00",
        "flow_end": "2026-07-11T06:00:00",
        "booked_capacity_kwh_h": 1000.0,
        "runtime_hours": 24.0,
        "tariff_eur_mwh_h": 1.0,
        "premium_eur_mwh_h": 0.0,
        "state": "Finished",
    }

    first = storage.upsert([row])
    second = storage.upsert([row])

    assert first["inserted"] == 1
    assert second["inserted"] == 0
    assert second["unchanged"] == 1


def test_excel_width_mapping_covers_columns_and_header_only_export(tmp_path) -> None:
    storage = AuctionStorage(tmp_path / "test.db")
    output = storage.export_excel(tmp_path / "result.xlsx")

    assert AuctionStorage.EXCEL_COLUMN_WIDTHS == APPROVED_EXCEL_COLUMN_WIDTHS
    assert tuple(APPROVED_EXCEL_COLUMN_WIDTHS) == AuctionStorage.EXCEL_COLUMNS
    assert AuctionStorage.validate_excel(output)
    workbook = load_workbook(output)
    sheet = workbook["Auctions"]
    assert sheet.max_row == 1
    assert {
        header: sheet.column_dimensions[get_column_letter(index)].width
        for index, header in enumerate(AuctionStorage.EXCEL_COLUMNS, start=1)
    } == APPROVED_EXCEL_COLUMN_WIDTHS
    workbook.close()


def test_populated_excel_export_has_production_widths(tmp_path) -> None:
    storage = AuctionStorage(tmp_path / "test.db")
    storage.upsert([{
        "auction_id": "123", "auction_date": "2026-07-10T06:00:00",
        "exit_market": "", "entry_market": "VGS Storage Hub",
        "direction": "entry", "network_point": "VGS Storage Hub (4290)",
        "network_point_id": "NP-1", "tso_exit": "", "tso_entry": "Test TSO",
        "product_type": "Day Ahead", "flow_start": "2026-07-10T06:00:00",
        "flow_end": "2026-07-11T06:00:00", "booked_capacity_kwh_h": 1000.0,
        "runtime_hours": 24.0, "tariff_eur_mwh_h": 1.0,
        "premium_eur_mwh_h": 0.0, "state": "Finished",
    }])

    output = storage.export_excel(tmp_path / "result.xlsx")

    assert AuctionStorage.validate_excel(output)
    assert pd.read_excel(output)["Auction ID"].astype(str).tolist() == ["123"]


def test_excel_validation_rejects_missing_or_incorrect_widths(tmp_path) -> None:
    path = tmp_path / "result.xlsx"
    pd.DataFrame(columns=AuctionStorage.EXCEL_COLUMNS).to_excel(
        path, index=False, sheet_name="Auctions"
    )
    assert not AuctionStorage.validate_excel(path)
    AuctionStorage.apply_excel_widths(path)
    workbook = load_workbook(path)
    workbook["Auctions"].column_dimensions["A"].width = 12
    workbook.save(path)
    workbook.close()

    assert not AuctionStorage.validate_excel(path)


def historical_row(**changes):
    row = {
        "auction_id": "123", "auction_date": "2026-07-10T06:00:00",
        "exit_market": "", "entry_market": "", "direction": "entry",
        "network_point": "VGS Storage Hub (4290)", "network_point_id": "NP-1",
        "tso_exit": "", "tso_entry": "Test TSO", "product_type": "Day Ahead",
        "flow_start": "2026-07-10T06:00:00", "flow_end": "2026-07-11T06:00:00",
        "booked_capacity_kwh_h": 1000.0, "runtime_hours": 24.0,
        "tariff_eur_mwh_h": 1.0, "premium_eur_mwh_h": 0.0, "state": "Finished",
    }
    row.update(changes)
    return row


def test_identical_rows_in_one_batch_preserve_idempotent_counts(tmp_path):
    storage = AuctionStorage(tmp_path / "test.db")
    row = historical_row()

    result = storage.upsert([row, dict(row)])

    assert result == {"processed": 2, "inserted": 1, "updated": 0, "unchanged": 1}


def test_conflicting_batch_fails_before_inserting_into_empty_database(tmp_path):
    storage = AuctionStorage(tmp_path / "test.db")
    row = historical_row()

    with pytest.raises(AuctionStorageError, match="conflicting rows"):
        storage.upsert([row, {**row, "state": "Open"}])

    with sqlite3.connect(storage.database_path) as connection:
        assert connection.execute("SELECT count(*) FROM auctions").fetchone()[0] == 0


def test_conflicting_batch_cannot_modify_existing_auction(tmp_path):
    storage = AuctionStorage(tmp_path / "test.db")
    original = historical_row(state="Original")
    storage.upsert([original])

    with pytest.raises(AuctionStorageError, match="conflicting rows"):
        storage.upsert([
            {**original, "state": "First change"},
            {**original, "state": "Second change"},
        ])

    with sqlite3.connect(storage.database_path) as connection:
        assert connection.execute("SELECT state FROM auctions").fetchone()[0] == "Original"


@pytest.mark.parametrize("network_point_id", ["", " \t "])
def test_blank_network_point_id_fails_before_storage_mutation(
    tmp_path, network_point_id
):
    storage = AuctionStorage(tmp_path / "test.db")
    original = historical_row()
    storage.upsert([original])

    with pytest.raises(AuctionStorageError, match="nonblank string"):
        storage.upsert([
            {**original, "state": "Changed"},
            historical_row(auction_id="blank", network_point_id=network_point_id),
        ])

    with sqlite3.connect(storage.database_path) as connection:
        auctions = connection.execute("SELECT state FROM auctions").fetchall()
    assert auctions == [("Finished",)]


def test_explicit_historical_backfill_is_idempotent_and_preserves_complete_values(tmp_path):
    storage = AuctionStorage(tmp_path / "test.db")
    storage.upsert([
        historical_row(),
        historical_row(auction_id="124", network_point_id="NP-2",
                       entry_market="VGS Storage Hub"),
    ])

    first = storage.backfill_historical_market_storage()
    second = storage.backfill_historical_market_storage()

    assert (first.examined, first.updated, first.unchanged, first.committed) == (2, 1, 1, True)
    assert (second.examined, second.updated, second.unchanged) == (2, 0, 2)
    persisted = storage.historical_market_storage_audit()
    assert len(persisted) == 4
    assert first.run_id != second.run_id
    assert len(storage.historical_market_storage_runs()) == 2
    assert [row["status"] for row in storage.historical_market_storage_runs()] == [
        "committed", "committed"
    ]
    assert [row["status"] for row in persisted[:2]] == ["updated", "unchanged/already_complete"]
    assert [item.status for item in second.audit] == [
        HistoricalBackfillStatus.ALREADY_COMPLETE,
        HistoricalBackfillStatus.ALREADY_COMPLETE,
    ]


def test_historical_backfill_conflict_unresolvable_invalid_and_partial_fill(tmp_path):
    storage = AuctionStorage(tmp_path / "test.db")
    storage.upsert([
        historical_row(auction_id="partial", network_point_id="1", exit_market="keep"),
        historical_row(auction_id="conflict", network_point_id="2", entry_market="Wrong"),
        historical_row(auction_id="unknown", network_point_id="3", network_point="Unknown"),
        historical_row(auction_id="invalid", network_point_id="4", direction="sideways"),
        historical_row(auction_id="bundle", network_point_id="5", direction="bundle"),
    ])

    summary = storage.backfill_historical_market_storage()

    assert (summary.examined, summary.updated, summary.unchanged, summary.skipped,
            summary.conflicts, summary.invalid) == (5, 1, 0, 2, 1, 1)
    assert [item.reason_code for item in summary.audit] == [
        "missing_values_filled", "reference_conflict", "reference_unresolvable",
        "invalid_historical_row", "insufficient_bundle_identity",
    ]
    with sqlite3.connect(storage.database_path) as connection:
        values = connection.execute(
            "SELECT auction_id, exit_market, entry_market FROM auctions ORDER BY id"
        ).fetchall()
    assert values == [
        ("partial", "keep", "VGS Storage Hub"),
        ("conflict", "", "Wrong"),
        ("unknown", "", ""),
        ("invalid", "", ""),
        ("bundle", "", ""),
    ]


def test_historical_backfill_mid_write_failure_rolls_back_rows_and_audit(tmp_path, monkeypatch):
    storage = AuctionStorage(tmp_path / "test.db")
    storage.upsert([historical_row(), historical_row(auction_id="124", network_point_id="2")])
    original = AuctionStorage._resolve_historical_market_storage
    calls = 0

    def fail_on_second(row, catalog, run_id, position):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise sqlite3.OperationalError("injected backfill failure")
        return original(row, catalog, run_id, position)

    monkeypatch.setattr(AuctionStorage, "_resolve_historical_market_storage", staticmethod(fail_on_second))
    with pytest.raises(sqlite3.OperationalError, match="injected"):
        storage.backfill_historical_market_storage()
    with sqlite3.connect(storage.database_path) as connection:
        assert connection.execute(
            "SELECT count(*) FROM auctions WHERE entry_market != ''"
        ).fetchone()[0] == 0
        assert connection.execute(
            "SELECT count(*) FROM historical_market_storage_audit"
        ).fetchone()[0] == 0
        assert connection.execute(
            "SELECT count(*) FROM historical_market_storage_runs"
        ).fetchone()[0] == 0


def test_historical_backfill_blank_and_canonical_equivalence(tmp_path):
    storage = AuctionStorage(tmp_path / "test.db")
    storage.upsert([
        historical_row(auction_id="blank", network_point_id="1", entry_market=" \t"),
        historical_row(auction_id="equivalent", network_point_id="2",
                       entry_market="  vGs StOrAgE HuB  "),
    ])
    summary = storage.backfill_historical_market_storage()
    assert (summary.updated, summary.unchanged, summary.conflicts) == (1, 1, 0)
    with sqlite3.connect(storage.database_path) as connection:
        values = connection.execute(
            "SELECT entry_market FROM auctions ORDER BY id"
        ).fetchall()
    assert values == [("VGS Storage Hub",), ("  vGs StOrAgE HuB  ",)]


@pytest.mark.parametrize("changes", [
    {"auction_date": "bad"},
    {"flow_start": "2026-07-12T06:00:00"},
    {"runtime_hours": float("inf")},
    {"product_type": ""},
])
def test_malformed_historical_rows_are_invalid_and_untouched(tmp_path, changes):
    storage = AuctionStorage(tmp_path / "test.db")
    storage.upsert([historical_row(**changes)])
    summary = storage.backfill_historical_market_storage()
    assert (summary.invalid, summary.updated) == (1, 0)
    assert summary.audit[0].reason_code == "invalid_historical_row"


def test_connections_enforce_audit_foreign_keys_and_restrict_delete(tmp_path):
    storage = AuctionStorage(tmp_path / "test.db")
    storage.upsert([historical_row()])
    storage.backfill_historical_market_storage()
    with storage._connect() as connection:
        assert connection.execute("PRAGMA foreign_keys").fetchone()[0] == 1
        with pytest.raises(sqlite3.IntegrityError):
            connection.execute("DELETE FROM auctions")
        with pytest.raises(sqlite3.IntegrityError):
            connection.execute("DELETE FROM historical_market_storage_runs")


def test_audit_auction_row_id_index_is_exact_and_idempotent(tmp_path):
    storage = AuctionStorage(tmp_path / "test.db")
    AuctionStorage(storage.database_path)
    with closing(storage._connect()) as connection:
        indexes = {row[1]: row for row in connection.execute(
            "PRAGMA index_list(historical_market_storage_audit)"
        )}
        assert AuctionStorage.AUDIT_ROW_INDEX in indexes
        assert [row[2] for row in connection.execute(
            f"PRAGMA index_info({AuctionStorage.AUDIT_ROW_INDEX})"
        )] == ["auction_row_id"]


def test_current_schema_reopens_without_mutation(tmp_path):
    path = tmp_path / "test.db"
    AuctionStorage(path)
    before = path.read_bytes()
    AuctionStorage(path)
    assert path.read_bytes() == before


def test_pre_p337_database_and_experimental_audit_schema_are_upgraded(tmp_path):
    path = tmp_path / "test.db"
    storage = AuctionStorage(path)
    with storage._connect() as connection:
        connection.execute("DROP TABLE historical_market_storage_audit")
        connection.execute("DROP TABLE historical_market_storage_runs")
        connection.execute("CREATE TABLE historical_market_storage_audit "
                           "(auction_row_id INTEGER PRIMARY KEY)")
    reopened = AuctionStorage(path)
    columns = {row["name"] for row in reopened._connect().execute(
        "PRAGMA table_info(historical_market_storage_audit)"
    )}
    assert {"run_id", "auction_row_id", "row_position"} <= columns


@pytest.mark.parametrize("definition", [
    "auction_row_id INTEGER PRIMARY KEY, extra TEXT",
    "auction_row_id INTEGER NOT NULL UNIQUE",
])
def test_unknown_experimental_like_schema_fails_closed(tmp_path, definition):
    path = tmp_path / "test.db"
    storage = AuctionStorage(path)
    with closing(storage._connect()) as connection, connection:
        connection.execute("DROP TABLE historical_market_storage_audit")
        connection.execute("DROP TABLE historical_market_storage_runs")
        connection.execute(f"CREATE TABLE historical_market_storage_audit ({definition})")
    before = path.read_bytes()
    with pytest.raises(AuctionStorageError, match="Unknown or partial"):
        AuctionStorage(path)
    assert path.read_bytes() == before


def test_experimental_migration_create_failure_restores_old_table(tmp_path, monkeypatch):
    path = tmp_path / "test.db"
    storage = AuctionStorage(path)
    with closing(storage._connect()) as connection, connection:
        connection.execute("DROP TABLE historical_market_storage_audit")
        connection.execute("DROP TABLE historical_market_storage_runs")
        connection.execute("CREATE TABLE historical_market_storage_audit "
                           "(auction_row_id INTEGER PRIMARY KEY)")
        connection.execute("INSERT INTO historical_market_storage_audit VALUES (17)")
    original = AuctionStorage._create_historical_tables.__func__

    def fail_after_first_create(cls, connection):
        connection.execute(cls.RUNS_SQL)
        raise sqlite3.OperationalError("injected migration failure")

    monkeypatch.setattr(AuctionStorage, "_create_historical_tables", classmethod(fail_after_first_create))
    with pytest.raises(sqlite3.OperationalError, match="injected"):
        AuctionStorage(path)
    monkeypatch.setattr(AuctionStorage, "_create_historical_tables", classmethod(original))
    with sqlite3.connect(path) as connection:
        assert connection.execute(
            "SELECT auction_row_id FROM historical_market_storage_audit"
        ).fetchall() == [(17,)]
        assert connection.execute(
            "SELECT count(*) FROM sqlite_master WHERE type='table' "
            "AND name='historical_market_storage_runs'"
        ).fetchone()[0] == 0


@pytest.mark.parametrize(("direction", "network_point", "field"), [
    ("entry", "VGS Storage Hub (4290)", "entry_market"),
    ("exit", "Kulata (BG)/Sidirokastron (GR)", "exit_market"),
])
def test_required_side_null_is_missing_and_audit_preserves_null(
    tmp_path, direction, network_point, field
):
    storage = AuctionStorage(tmp_path / "test.db")
    storage.upsert([historical_row(direction=direction, network_point=network_point)])
    with sqlite3.connect(storage.database_path) as connection:
        connection.execute("PRAGMA foreign_keys=OFF")
        sql = connection.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='auctions'"
        ).fetchone()[0]
        connection.execute("DROP TABLE historical_market_storage_audit")
        connection.execute("DROP TABLE historical_market_storage_runs")
        connection.execute("ALTER TABLE auctions RENAME TO auctions_strict")
        connection.execute(sql.replace("exit_market TEXT NOT NULL", "exit_market TEXT")
                           .replace("entry_market TEXT NOT NULL", "entry_market TEXT"))
        columns = [row[1] for row in connection.execute("PRAGMA table_info(auctions)")]
        connection.execute(
            f"INSERT INTO auctions ({', '.join(columns)}) "
            f"SELECT {', '.join(columns)} FROM auctions_strict"
        )
        connection.execute("DROP TABLE auctions_strict")
        connection.execute(f"UPDATE auctions SET {field}=NULL")
        AuctionStorage._create_historical_tables(connection)
    summary = storage.backfill_historical_market_storage()
    assert (summary.updated, summary.conflicts) == (1, 0)
    audit_field = f"previous_{field}"
    assert getattr(summary.audit[0], audit_field) is None
    assert storage.historical_market_storage_audit()[0][audit_field] is None


@pytest.mark.parametrize(("start", "end", "invalid"), [
    ("2026-07-10T06:00:00", "2026-07-11T06:00:00+00:00", True),
    ("2026-07-10T06:00:00+02:00", "2026-07-10T05:00:00+00:00", False),
    ("2026-07-10T06:00:00+00:00", "2026-07-10T07:00:00+02:00", True),
])
def test_timezone_awareness_and_instant_order(start, end, invalid, tmp_path):
    storage = AuctionStorage(tmp_path / "test.db")
    storage.upsert([historical_row(flow_start=start, flow_end=end)])
    summary = storage.backfill_historical_market_storage()
    assert summary.invalid == int(invalid)
    assert summary.updated == int(not invalid)


def test_connections_close_on_success_and_exception(tmp_path, monkeypatch):
    storage = AuctionStorage(tmp_path / "test.db")
    closed = []

    class TrackedConnection(sqlite3.Connection):
        def close(self):
            closed.append(True)
            super().close()

    monkeypatch.setattr(storage, "_connect", lambda: sqlite3.connect(
        storage.database_path, factory=TrackedConnection
    ))
    storage.operations()
    assert closed == [True]
    monkeypatch.setattr(AuctionStorage, "_resolve_historical_market_storage",
                        staticmethod(lambda *args: (_ for _ in ()).throw(RuntimeError("boom"))))
    storage.upsert([historical_row()])
    with pytest.raises(RuntimeError, match="boom"):
        storage.backfill_historical_market_storage()
    assert closed[-1] is True


def test_connection_close_failure_is_visible_after_success(tmp_path, monkeypatch):
    storage = AuctionStorage(tmp_path / "test.db")

    class CloseFailure(sqlite3.Connection):
        def close(self):
            super().close()
            raise RuntimeError("injected close failure")

    monkeypatch.setattr(storage, "_connect", lambda: sqlite3.connect(
        storage.database_path, factory=CloseFailure
    ))
    with pytest.raises(RuntimeError, match="close failure"):
        storage.operations()


def test_processing_failure_remains_primary_when_close_fails(tmp_path, monkeypatch):
    storage = AuctionStorage(tmp_path / "test.db")
    storage.upsert([historical_row()])
    close_calls = 0

    class CloseFailure(sqlite3.Connection):
        def close(self):
            nonlocal close_calls
            close_calls += 1
            super().close()
            raise RuntimeError("injected close failure")

    def connect():
        connection = sqlite3.connect(storage.database_path, factory=CloseFailure)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys=ON")
        return connection

    monkeypatch.setattr(storage, "_connect", connect)
    monkeypatch.setattr(AuctionStorage, "_resolve_historical_market_storage",
                        staticmethod(lambda *args: (_ for _ in ()).throw(ValueError("processing"))))
    with pytest.raises(ValueError, match="processing") as caught:
        storage.backfill_historical_market_storage()
    assert close_calls == 1
    assert any("close failure" in note for note in caught.value.__notes__)


def test_processing_primary_survives_close_failure_and_broken_add_note(
    tmp_path, monkeypatch
):
    storage = AuctionStorage(tmp_path / "test.db")
    storage.upsert([historical_row()])
    close_calls = 0

    class BrokenDiagnostic(BaseException):
        def add_note(self, note):
            raise RuntimeError("injected add_note failure")

    primary = BrokenDiagnostic("primary processing failure")

    class CloseFailure(sqlite3.Connection):
        def close(self):
            nonlocal close_calls
            close_calls += 1
            super().close()
            raise RuntimeError("injected close failure")

    def connect():
        connection = sqlite3.connect(storage.database_path, factory=CloseFailure)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys=ON")
        return connection

    def fail_processing(*args):
        raise primary

    monkeypatch.setattr(storage, "_connect", connect)
    monkeypatch.setattr(
        AuctionStorage, "_resolve_historical_market_storage",
        staticmethod(fail_processing),
    )
    with pytest.raises(BrokenDiagnostic, match="primary processing failure") as caught:
        storage.backfill_historical_market_storage()
    assert caught.value is primary
    assert close_calls == 1
    assert caught.value.__traceback__ is not None


def test_commit_failure_returns_no_summary_and_rolls_back(tmp_path, monkeypatch):
    storage = AuctionStorage(tmp_path / "test.db")
    storage.upsert([historical_row()])

    close_calls = 0

    class CommitFailure(sqlite3.Connection):
        def commit(self):
            raise sqlite3.OperationalError("injected commit failure")

        def close(self):
            nonlocal close_calls
            close_calls += 1
            super().close()
            raise RuntimeError("injected close failure")

    def connect():
        connection = sqlite3.connect(storage.database_path, factory=CommitFailure)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys=ON")
        return connection

    monkeypatch.setattr(storage, "_connect", connect)
    with pytest.raises(sqlite3.OperationalError, match="commit failure") as caught:
        storage.backfill_historical_market_storage()
    assert close_calls == 1
    assert any("close failure" in note for note in caught.value.__notes__)
    with sqlite3.connect(storage.database_path) as connection:
        assert connection.execute("SELECT entry_market FROM auctions").fetchone()[0] == ""
        assert connection.execute(
            "SELECT count(*) FROM historical_market_storage_runs"
        ).fetchone()[0] == 0


def test_commit_primary_survives_rollback_close_and_broken_add_note(
    tmp_path, monkeypatch
):
    storage = AuctionStorage(tmp_path / "test.db")
    storage.upsert([historical_row()])
    commit_calls = rollback_calls = close_calls = 0

    class BrokenDiagnostic(BaseException):
        def add_note(self, note):
            raise RuntimeError("injected add_note failure")

    primary = BrokenDiagnostic("primary commit failure")

    class TripleFailure(sqlite3.Connection):
        def commit(self):
            nonlocal commit_calls
            commit_calls += 1
            raise primary

        def rollback(self):
            nonlocal rollback_calls
            rollback_calls += 1
            raise RuntimeError("injected rollback failure")

        def close(self):
            nonlocal close_calls
            close_calls += 1
            super().close()
            raise RuntimeError("injected close failure")

    def connect():
        connection = sqlite3.connect(storage.database_path, factory=TripleFailure)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys=ON")
        return connection

    monkeypatch.setattr(storage, "_connect", connect)
    with pytest.raises(BrokenDiagnostic, match="primary commit failure") as caught:
        storage.backfill_historical_market_storage()
    assert caught.value is primary
    assert (commit_calls, rollback_calls, close_calls) == (1, 1, 1)
    assert caught.value.__traceback__ is not None
    with sqlite3.connect(storage.database_path) as connection:
        assert connection.execute("SELECT entry_market FROM auctions").fetchone()[0] == ""
        assert connection.execute(
            "SELECT count(*) FROM historical_market_storage_runs"
        ).fetchone()[0] == 0
        assert connection.execute(
            "SELECT count(*) FROM historical_market_storage_audit"
        ).fetchone()[0] == 0
        assert connection.execute(
            "SELECT count(*) FROM historical_market_storage_audit"
        ).fetchone()[0] == 0


def test_begin_immediate_prevents_concurrent_lost_update(tmp_path, monkeypatch):
    storage = AuctionStorage(tmp_path / "test.db")
    storage.upsert([historical_row()])
    reservation_held = threading.Event()
    release = threading.Event()
    original = AuctionStorage._resolve_historical_market_storage

    def pause_after_select(row, catalog, run_id, position):
        reservation_held.set()
        assert release.wait(5)
        return original(row, catalog, run_id, position)

    monkeypatch.setattr(AuctionStorage, "_resolve_historical_market_storage",
                        staticmethod(pause_after_select))
    errors = []
    worker = threading.Thread(
        target=lambda: storage.backfill_historical_market_storage(), daemon=True
    )
    worker.start()
    assert reservation_held.wait(5)
    try:
        concurrent = sqlite3.connect(storage.database_path, timeout=0)
        with pytest.raises(sqlite3.OperationalError, match="locked"):
            concurrent.execute("UPDATE auctions SET entry_market='raced'")
        concurrent.close()
    finally:
        release.set()
        worker.join(5)
    assert not worker.is_alive()
    assert not errors
    with sqlite3.connect(storage.database_path) as connection:
        assert connection.execute(
            "SELECT entry_market FROM auctions"
        ).fetchone()[0] == "VGS Storage Hub"


def test_concurrent_schema_initializers_serialize_before_fingerprint(tmp_path, monkeypatch):
    path = tmp_path / "test.db"
    storage = AuctionStorage(path)
    with sqlite3.connect(path) as connection:
        connection.execute("DROP TABLE historical_market_storage_audit")
        connection.execute("DROP TABLE historical_market_storage_runs")
        connection.execute("CREATE TABLE historical_market_storage_audit "
                           "(auction_row_id INTEGER PRIMARY KEY)")
        connection.execute("INSERT INTO historical_market_storage_audit VALUES (17)")

    first_reserved = threading.Event()
    release_first = threading.Event()
    second_attempting_reservation = threading.Event()
    second_reserved = threading.Event()

    class CoordinatedConnection(sqlite3.Connection):
        def execute(self, sql, parameters=(), /):
            if (
                threading.current_thread().name == "second-initializer"
                and sql.strip().upper() == "BEGIN IMMEDIATE"
            ):
                second_attempting_reservation.set()
            return super().execute(sql, parameters)

    def connect(instance):
        connection = sqlite3.connect(
            instance.database_path, factory=CoordinatedConnection
        )
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys=ON")
        if connection.execute("PRAGMA foreign_keys").fetchone()[0] != 1:
            connection.close()
            raise AuctionStorageError(
                "SQLite foreign key enforcement could not be enabled."
            )
        return connection

    def after_reservation(connection):
        if threading.current_thread().name == "first-initializer":
            first_reserved.set()
            assert release_first.wait(5)
        else:
            second_reserved.set()

    monkeypatch.setattr(AuctionStorage, "_connect", connect)
    monkeypatch.setattr(AuctionStorage, "_after_schema_reservation", staticmethod(after_reservation))
    errors = []

    def initialize():
        try:
            AuctionStorage(path)
        except BaseException as exc:
            errors.append(exc)

    first = threading.Thread(target=initialize, name="first-initializer")
    second = threading.Thread(target=initialize, name="second-initializer")
    first.start()
    assert first_reserved.wait(5)
    second.start()
    assert second_attempting_reservation.wait(5)
    assert not second_reserved.is_set()
    release_first.set()
    first.join(5)
    second.join(5)
    assert not errors
    assert second_reserved.is_set()
    reopened = AuctionStorage(path)
    assert reopened.historical_market_storage_runs() == []
    with sqlite3.connect(path) as connection:
        assert connection.execute("SELECT count(*) FROM auctions").fetchone()[0] == 0
        assert connection.execute(
            "SELECT count(*) FROM sqlite_master WHERE type='table' "
            "AND name IN ('historical_market_storage_runs', "
            "'historical_market_storage_audit')"
        ).fetchone()[0] == 2


def test_schema_lock_timeout_preserves_busy_error_and_closes_once(tmp_path, monkeypatch):
    path = tmp_path / "test.db"
    AuctionStorage(path)
    locker = sqlite3.connect(path)
    locker.execute("BEGIN IMMEDIATE")
    close_calls = 0

    class BusyConnection(sqlite3.Connection):
        def close(self):
            nonlocal close_calls
            close_calls += 1
            super().close()
            raise RuntimeError("injected close failure")

    storage = object.__new__(AuctionStorage)
    storage.database_path = path

    def connect():
        connection = sqlite3.connect(path, timeout=0.01, factory=BusyConnection)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys=ON")
        return connection

    monkeypatch.setattr(storage, "_connect", connect)
    try:
        with pytest.raises(sqlite3.OperationalError, match="locked") as caught:
            storage._create_schema()
    finally:
        locker.rollback()
        locker.close()
    assert close_calls == 1
    assert any("close failure" in note for note in caught.value.__notes__)
    with sqlite3.connect(path) as connection:
        assert connection.execute(
            "SELECT count(*) FROM sqlite_master WHERE type='table' "
            "AND name IN ('historical_market_storage_runs', "
            "'historical_market_storage_audit')"
        ).fetchone()[0] == 2


def test_migration_failure_remains_primary_when_close_fails(tmp_path, monkeypatch):
    path = tmp_path / "test.db"
    AuctionStorage(path)
    with sqlite3.connect(path) as connection:
        connection.execute("DROP TABLE historical_market_storage_audit")
        connection.execute("DROP TABLE historical_market_storage_runs")
        connection.execute("CREATE TABLE historical_market_storage_audit "
                           "(auction_row_id INTEGER PRIMARY KEY)")
        connection.execute("INSERT INTO historical_market_storage_audit VALUES (17)")
    close_calls = 0

    class MigrationConnection(sqlite3.Connection):
        def close(self):
            nonlocal close_calls
            close_calls += 1
            super().close()
            raise RuntimeError("injected close failure")

    storage = object.__new__(AuctionStorage)
    storage.database_path = path

    def connect():
        connection = sqlite3.connect(path, factory=MigrationConnection)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys=ON")
        return connection

    def fail_create(cls, connection):
        connection.execute(cls.RUNS_SQL)
        raise sqlite3.OperationalError("injected migration failure")

    monkeypatch.setattr(storage, "_connect", connect)
    monkeypatch.setattr(AuctionStorage, "_create_historical_tables", classmethod(fail_create))
    with pytest.raises(sqlite3.OperationalError, match="migration failure") as caught:
        storage._create_schema()
    assert close_calls == 1
    assert any("close failure" in note for note in caught.value.__notes__)
    with sqlite3.connect(path) as connection:
        assert connection.execute(
            "SELECT auction_row_id FROM historical_market_storage_audit"
        ).fetchall() == [(17,)]
        assert connection.execute(
            "SELECT count(*) FROM sqlite_master WHERE type='table' "
            "AND name='historical_market_storage_runs'"
        ).fetchone()[0] == 0
