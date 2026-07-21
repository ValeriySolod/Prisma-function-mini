import os
from datetime import date, datetime
from decimal import Decimal

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from mini_domain import (
    CapacityType, MiniDateRange, NormalizedAuctionRecord, OUTPUT_COLUMNS,
    ProductType, SourceImportRequest,
)
from mini_storage import MiniAuctionStorage
from mini_workbook import MiniWorkbookError, MiniWorkbookPublisher
from runtime_paths import RuntimePaths


def paths(root):
    app_root = root / "PrismaFunctionMini"
    return RuntimePaths(
        app_root, app_root / "data" / "prisma_function_mini.db",
        app_root / "data" / "result" / "prisma_function_mini.xlsx",
        app_root / "state" / "prisma_function_mini_state.json",
        app_root / "logs" / "prisma-function-mini.log", app_root / "temporary-downloads",
    )


def request(sha="a" * 64):
    selected = MiniDateRange(date(2026, 7, 1), date(2026, 7, 5))
    return SourceImportRequest(selected, "source.csv", sha, 10)


def record(number, **changes):
    values = dict(
        auction_id=f"A-{number}", network_point_id=f"NP-{number}",
        auction_date=date(2026, 7, number), exit_market_or_storage="Exit Market",
        entry_market_or_storage="Entry Storage", capacity_type=CapacityType.BUNDLE,
        network_point=f"Point {number}", product_type=ProductType.DAY_AHEAD,
        flow_start=datetime(2026, 7, number + 1, 6),
        flow_end=datetime(2026, 7, number + 2, 6),
        booked_capacity_kwh_h=Decimal("1234.5"), duration_hours=Decimal("24"),
        auction_tariff_eur_mwh_h=Decimal("1.25"),
    )
    values.update(changes)
    return NormalizedAuctionRecord(**values)


def test_publication_uses_authoritative_history_order_types_widths_and_formats(tmp_path):
    storage = MiniAuctionStorage(paths=paths(tmp_path))
    storage.store(request(), [record(3), record(1), record(2)])

    output = MiniWorkbookPublisher(storage).publish()

    workbook = load_workbook(output, data_only=True)
    sheet = workbook["Auctions"]
    assert tuple(cell.value for cell in sheet[1]) == OUTPUT_COLUMNS
    assert [sheet.cell(row, 5).value for row in range(2, 5)] == [
        "Point 1", "Point 2", "Point 3",
    ]
    assert isinstance(sheet["A2"].value, datetime)
    assert isinstance(sheet["G2"].value, datetime)
    assert isinstance(sheet["I2"].value, (int, float))
    assert sheet["I2"].value == 1234.5
    assert {
        header: sheet.column_dimensions[get_column_letter(index)].width
        for index, header in enumerate(OUTPUT_COLUMNS, start=1)
    } == MiniWorkbookPublisher.COLUMN_WIDTHS
    assert sheet["A2"].number_format == "yyyy-mm-dd"
    assert sheet["G2"].number_format == "yyyy-mm-dd hh:mm"
    workbook.close()


def test_cumulative_publication_preserves_rows_and_exact_retry_is_unchanged(tmp_path):
    storage = MiniAuctionStorage(paths=paths(tmp_path))
    publisher = MiniWorkbookPublisher(storage)
    storage.store(request(), [record(1)])
    output = publisher.publish()
    first_bytes = output.read_bytes()
    first_mtime = output.stat().st_mtime_ns

    storage.store(request(), [record(1)])
    assert publisher.publish().read_bytes() == first_bytes
    assert output.stat().st_mtime_ns == first_mtime

    storage.store(request("b" * 64), [record(2)])
    publisher.publish()
    workbook = load_workbook(output, data_only=True)
    assert [cell.value for cell in workbook["Auctions"]["E"][1:]] == ["Point 1", "Point 2"]
    workbook.close()


def test_failed_atomic_replace_preserves_last_valid_workbook(tmp_path, monkeypatch):
    storage = MiniAuctionStorage(paths=paths(tmp_path))
    publisher = MiniWorkbookPublisher(storage)
    storage.store(request(), [record(1)])
    output = publisher.publish()
    previous = output.read_bytes()
    storage.store(request("b" * 64), [record(2)])

    def fail_replace(_source, _destination):
        raise PermissionError("locked")

    monkeypatch.setattr(os, "replace", fail_replace)
    with pytest.raises(MiniWorkbookError, match="open or locked"):
        publisher.publish()

    assert output.read_bytes() == previous
    assert not list(output.parent.glob(f".{output.stem}-*.xlsx"))


def test_staging_or_validation_failure_leaves_existing_workbook_unchanged(tmp_path, monkeypatch):
    storage = MiniAuctionStorage(paths=paths(tmp_path))
    publisher = MiniWorkbookPublisher(storage)
    storage.store(request(), [record(1)])
    output = publisher.publish()
    previous = output.read_bytes()
    storage.store(request("b" * 64), [record(2)])
    monkeypatch.setattr(publisher, "validate", lambda *_args: False)

    with pytest.raises(MiniWorkbookError, match="failed validation"):
        publisher.publish()

    assert output.read_bytes() == previous
    assert not list(output.parent.glob(f".{output.stem}-*.xlsx"))


def test_empty_history_publishes_valid_header_only_workbook(tmp_path):
    publisher = MiniWorkbookPublisher(MiniAuctionStorage(paths=paths(tmp_path)))
    output = publisher.publish()
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["Auctions"]
    assert workbook["Auctions"].max_row == 1
    workbook.close()
