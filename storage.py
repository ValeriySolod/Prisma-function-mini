from __future__ import annotations

import json
import math
import os
import sqlite3
import tempfile
import uuid
from contextlib import closing, contextmanager
from dataclasses import dataclass
from datetime import datetime, timezone
from enum import Enum
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from prisma_references import (
    DEFAULT_PRISMA_REFERENCES,
    PrismaReferenceCatalog,
    ReferenceSide,
)


class AuctionStorageError(RuntimeError):
    pass


class HistoricalBackfillStatus(str, Enum):
    UPDATED = "updated"
    ALREADY_COMPLETE = "unchanged/already_complete"
    UNRESOLVABLE = "skipped/unresolvable"
    CONFLICT = "conflict"
    INVALID = "invalid"


@dataclass(frozen=True)
class HistoricalBackfillAudit:
    run_id: str
    auction_row_id: int
    row_position: int
    auction_id: str
    row_key: str
    previous_exit_market: str | None
    previous_entry_market: str | None
    proposed_exit_market: str | None
    proposed_entry_market: str | None
    final_exit_market: str | None
    final_entry_market: str | None
    status: HistoricalBackfillStatus
    reason_code: str
    message: str
    changed: bool


@dataclass(frozen=True)
class HistoricalBackfillSummary:
    run_id: str
    examined: int
    updated: int
    unchanged: int
    skipped: int
    conflicts: int
    invalid: int
    committed: bool
    audit: tuple[HistoricalBackfillAudit, ...]


class AuctionStorage:
    AUCTION_IDENTITY_FIELDS = (
        "auction_id", "network_point_id", "direction", "flow_start", "flow_end",
    )
    AUCTION_PERSISTED_FIELDS = (
        "auction_id", "auction_date", "exit_market", "entry_market", "direction",
        "network_point", "network_point_id", "tso_exit", "tso_entry", "product_type",
        "flow_start", "flow_end", "booked_capacity_kwh_h", "runtime_hours",
        "tariff_eur_mwh_h", "premium_eur_mwh_h", "state",
    )
    AUDIT_ROW_INDEX = "idx_historical_market_storage_audit_auction_row_id"
    EXPERIMENTAL_AUDIT_COLUMNS = (
        ("auction_row_id", "INTEGER", 0, None, 1),
    )
    RUNS_SQL = """CREATE TABLE historical_market_storage_runs (
        run_id TEXT PRIMARY KEY,
        run_timestamp_utc TEXT NOT NULL,
        examined INTEGER NOT NULL, updated INTEGER NOT NULL,
        unchanged INTEGER NOT NULL, skipped INTEGER NOT NULL,
        conflicts INTEGER NOT NULL, invalid INTEGER NOT NULL,
        status TEXT NOT NULL CHECK(status = 'committed'),
        CHECK(examined = updated + unchanged + skipped + conflicts + invalid)
    )"""
    AUDIT_SQL = """CREATE TABLE historical_market_storage_audit (
        run_id TEXT NOT NULL,
        auction_row_id INTEGER NOT NULL,
        row_position INTEGER NOT NULL,
        auction_id TEXT NOT NULL,
        row_key TEXT NOT NULL,
        previous_exit_market TEXT,
        previous_entry_market TEXT,
        proposed_exit_market TEXT,
        proposed_entry_market TEXT,
        final_exit_market TEXT,
        final_entry_market TEXT,
        status TEXT NOT NULL,
        reason_code TEXT NOT NULL,
        message TEXT NOT NULL,
        changed INTEGER NOT NULL CHECK(changed IN (0, 1)),
        PRIMARY KEY(run_id, auction_row_id),
        UNIQUE(run_id, row_position),
        FOREIGN KEY(run_id) REFERENCES historical_market_storage_runs(run_id)
            ON DELETE RESTRICT,
        FOREIGN KEY(auction_row_id) REFERENCES auctions(id)
            ON DELETE RESTRICT
    )"""
    EXCEL_COLUMNS = (
        "Auction Date", "Exit Market/Storage", "Entry Market/Storage",
        "Capacity Type", "Network Point Name", "Product Type", "Flow Start",
        "Flow End", "Booked Capacity, kWh/h", "Runtime Hours",
        "Tariff, EUR/MWh/h", "Premium, EUR/MWh/h", "Auction ID",
        "TSO Exit", "TSO Entry", "Status",
    )
    EXCEL_COLUMN_WIDTHS = {
        "Auction Date": 21,
        "Exit Market/Storage": 22,
        "Entry Market/Storage": 22,
        "Capacity Type": 15,
        "Network Point Name": 36,
        "Product Type": 14,
        "Flow Start": 21,
        "Flow End": 21,
        "Booked Capacity, kWh/h": 24,
        "Runtime Hours": 15,
        "Tariff, EUR/MWh/h": 20,
        "Premium, EUR/MWh/h": 21,
        "Auction ID": 16,
        "TSO Exit": 30,
        "TSO Entry": 30,
        "Status": 14,
    }
    EXCEL_WIDTH_TOLERANCE = 1e-6
    def __init__(self, database_path: Path) -> None:
        database_path.parent.mkdir(parents=True, exist_ok=True)
        self.database_path = database_path
        self._create_schema()

    def _connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self.database_path)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys = ON")
        if connection.execute("PRAGMA foreign_keys").fetchone()[0] != 1:
            primary = AuctionStorageError(
                "SQLite foreign key enforcement could not be enabled."
            )
            try:
                connection.close()
            except BaseException as close_error:
                self._add_exception_note(
                    primary, f"Connection close also failed: {close_error!r}"
                )
            raise primary
        return connection

    @staticmethod
    def _add_exception_note(primary: BaseException, diagnostic: str) -> None:
        """Attach failure context without ever replacing the primary exception."""
        try:
            add_note = getattr(primary, "add_note", None)
            if callable(add_note):
                add_note(diagnostic)
        except BaseException:
            pass

    @contextmanager
    def _connection(self):
        """Close one production connection without masking an active failure."""
        connection = self._connect()
        try:
            yield connection
        except BaseException as primary:
            try:
                connection.close()
            except BaseException as close_error:
                self._add_exception_note(
                    primary, f"Connection close also failed: {close_error!r}"
                )
            raise
        else:
            connection.close()

    @staticmethod
    @contextmanager
    def _transaction(connection: sqlite3.Connection):
        try:
            yield connection
            connection.commit()
        except BaseException as primary:
            try:
                connection.rollback()
            except BaseException as rollback_error:
                AuctionStorage._add_exception_note(
                    primary, f"Transaction rollback also failed: {rollback_error!r}"
                )
            raise

    @staticmethod
    def _after_schema_reservation(connection: sqlite3.Connection) -> None:
        """Private coordination seam used after the initialization write reservation."""

    def _create_schema(self) -> None:
        with self._connection() as connection, self._transaction(connection):
            connection.execute("BEGIN IMMEDIATE")
            self._after_schema_reservation(connection)
            for statement in (
                    """
                CREATE TABLE IF NOT EXISTS auctions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    auction_id TEXT NOT NULL, auction_date TEXT NOT NULL,
                    exit_market TEXT NOT NULL DEFAULT '', entry_market TEXT NOT NULL DEFAULT '',
                    direction TEXT NOT NULL, network_point TEXT NOT NULL,
                    network_point_id TEXT NOT NULL DEFAULT '', tso_exit TEXT NOT NULL DEFAULT '',
                    tso_entry TEXT NOT NULL DEFAULT '', product_type TEXT NOT NULL,
                    flow_start TEXT NOT NULL, flow_end TEXT NOT NULL,
                    booked_capacity_kwh_h REAL NOT NULL, runtime_hours REAL NOT NULL,
                    tariff_eur_mwh_h REAL NOT NULL, premium_eur_mwh_h REAL NOT NULL,
                    state TEXT NOT NULL DEFAULT '', created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE (auction_id, network_point_id, direction, flow_start, flow_end)
                )""",
                    """
                CREATE TABLE IF NOT EXISTS prisma_source_operations (
                    operation_id TEXT PRIMARY KEY,
                    source_date TEXT NOT NULL,
                    source_name TEXT NOT NULL,
                    sha256 TEXT NOT NULL,
                    status TEXT NOT NULL CHECK(status IN ('pending','data_committed','accepted')),
                    summary_json TEXT,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(source_date)
                )""",
            ):
                connection.execute(statement)
            self._ensure_historical_schema(connection)

    @staticmethod
    def _table_columns(connection: sqlite3.Connection, table: str) -> tuple[tuple[Any, ...], ...]:
        return tuple(
            (row[1], row[2], row[3], row[4], row[5])
            for row in connection.execute(f"PRAGMA table_info({table})")
        )

    @staticmethod
    def _foreign_keys(connection: sqlite3.Connection, table: str) -> tuple[tuple[Any, ...], ...]:
        return tuple(sorted(tuple(row[2:8]) for row in connection.execute(
            f"PRAGMA foreign_key_list({table})"
        )))

    @staticmethod
    def _indexes(connection: sqlite3.Connection, table: str) -> tuple[tuple[Any, ...], ...]:
        result = []
        for row in connection.execute(f"PRAGMA index_list({table})"):
            columns = tuple(item[2] for item in connection.execute(
                f"PRAGMA index_info({row[1]})"
            ))
            result.append((row[1], row[2], row[3], columns))
        return tuple(sorted(result))

    @classmethod
    def _schema_fingerprint(cls, connection: sqlite3.Connection) -> tuple[Any, ...]:
        return (
            cls._table_columns(connection, "historical_market_storage_runs"),
            cls._foreign_keys(connection, "historical_market_storage_runs"),
            cls._indexes(connection, "historical_market_storage_runs"),
            cls._table_columns(connection, "historical_market_storage_audit"),
            cls._foreign_keys(connection, "historical_market_storage_audit"),
            cls._indexes(connection, "historical_market_storage_audit"),
        )

    @classmethod
    def _create_historical_tables(cls, connection: sqlite3.Connection) -> None:
        connection.execute(cls.RUNS_SQL)
        connection.execute(cls.AUDIT_SQL)
        connection.execute(
            f"CREATE INDEX {cls.AUDIT_ROW_INDEX} "
            "ON historical_market_storage_audit(auction_row_id)"
        )

    @classmethod
    def _expected_historical_fingerprint(cls) -> tuple[Any, ...]:
        with closing(sqlite3.connect(":memory:")) as expected:
            expected.execute("CREATE TABLE auctions (id INTEGER PRIMARY KEY)")
            cls._create_historical_tables(expected)
            return cls._schema_fingerprint(expected)

    @classmethod
    def _ensure_historical_schema(cls, connection: sqlite3.Connection) -> None:
        audit_exists = bool(cls._table_columns(connection, "historical_market_storage_audit"))
        runs_exists = bool(cls._table_columns(connection, "historical_market_storage_runs"))
        if not audit_exists and not runs_exists:
            cls._create_historical_tables(connection)
            return
        experimental = (
            not runs_exists
            and cls._table_columns(connection, "historical_market_storage_audit")
            == cls.EXPERIMENTAL_AUDIT_COLUMNS
            and not cls._foreign_keys(connection, "historical_market_storage_audit")
            and not cls._indexes(connection, "historical_market_storage_audit")
        )
        if experimental:
            connection.execute("DROP TABLE historical_market_storage_audit")
            cls._create_historical_tables(connection)
            return
        if cls._schema_fingerprint(connection) != cls._expected_historical_fingerprint():
            raise AuctionStorageError(
                "Unknown or partial historical Market / Storage schema; database unchanged."
            )

    def historical_market_storage_audit(self) -> list[sqlite3.Row]:
        with self._connection() as connection:
            return list(connection.execute(
                "SELECT a.* FROM historical_market_storage_audit AS a "
                "JOIN historical_market_storage_runs AS r USING(run_id) "
                "ORDER BY r.rowid, a.row_position"
            ))

    def historical_market_storage_runs(self) -> list[sqlite3.Row]:
        with self._connection() as connection:
            return list(connection.execute(
                "SELECT * FROM historical_market_storage_runs ORDER BY rowid"
            ))

    def backfill_historical_market_storage(
        self,
        catalog: PrismaReferenceCatalog = DEFAULT_PRISMA_REFERENCES,
    ) -> HistoricalBackfillSummary:
        """Explicitly backfill missing historical Market / Storage values atomically."""
        run_id = uuid.uuid4().hex
        run_timestamp = datetime.now(timezone.utc).isoformat(timespec="microseconds")
        audit: list[HistoricalBackfillAudit] = []
        with self._connection() as connection, self._transaction(connection):
            connection.execute("BEGIN IMMEDIATE")
            rows = connection.execute("SELECT * FROM auctions ORDER BY id").fetchall()
            for position, row in enumerate(rows, start=1):
                item = self._resolve_historical_market_storage(
                    row, catalog, run_id, position
                )
                audit.append(item)
            counts = self._historical_backfill_counts(audit)
            connection.execute(
                "INSERT INTO historical_market_storage_runs "
                "(run_id, run_timestamp_utc, examined, updated, unchanged, skipped, "
                "conflicts, invalid, status) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'committed')",
                (run_id, run_timestamp, len(audit), *counts),
            )
            for item in audit:
                if item.changed:
                    connection.execute(
                        "UPDATE auctions SET exit_market=?, entry_market=?, "
                        "updated_at=CURRENT_TIMESTAMP WHERE id=?",
                        (item.final_exit_market, item.final_entry_market,
                         item.auction_row_id),
                    )
                connection.execute(
                    "INSERT INTO historical_market_storage_audit "
                    "(run_id, auction_row_id, row_position, auction_id, row_key, "
                    "previous_exit_market, previous_entry_market, proposed_exit_market, "
                    "proposed_entry_market, final_exit_market, final_entry_market, "
                    "status, reason_code, message, changed) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (
                        item.run_id, item.auction_row_id, item.row_position,
                        item.auction_id, item.row_key, item.previous_exit_market,
                        item.previous_entry_market, item.proposed_exit_market,
                        item.proposed_entry_market, item.final_exit_market,
                        item.final_entry_market, item.status.value, item.reason_code,
                        item.message, int(item.changed),
                    ),
                )
            self._validate_historical_backfill(connection, audit, run_id)
        counts_by_status = {
            HistoricalBackfillStatus.UPDATED: 0,
            HistoricalBackfillStatus.ALREADY_COMPLETE: 0,
            HistoricalBackfillStatus.UNRESOLVABLE: 0,
            HistoricalBackfillStatus.CONFLICT: 0,
            HistoricalBackfillStatus.INVALID: 0,
        }
        for item in audit:
            counts_by_status[item.status] += 1
        return HistoricalBackfillSummary(
            run_id, len(audit), counts_by_status[HistoricalBackfillStatus.UPDATED],
            counts_by_status[HistoricalBackfillStatus.ALREADY_COMPLETE],
            counts_by_status[HistoricalBackfillStatus.UNRESOLVABLE],
            counts_by_status[HistoricalBackfillStatus.CONFLICT],
            counts_by_status[HistoricalBackfillStatus.INVALID], True, tuple(audit),
        )

    @staticmethod
    def _historical_backfill_counts(audit: list[HistoricalBackfillAudit]) -> tuple[int, ...]:
        return tuple(sum(item.status is status for item in audit) for status in (
            HistoricalBackfillStatus.UPDATED, HistoricalBackfillStatus.ALREADY_COMPLETE,
            HistoricalBackfillStatus.UNRESOLVABLE, HistoricalBackfillStatus.CONFLICT,
            HistoricalBackfillStatus.INVALID,
        ))

    @staticmethod
    def _resolve_historical_market_storage(
        row: sqlite3.Row, catalog: PrismaReferenceCatalog, run_id: str, row_position: int
    ) -> HistoricalBackfillAudit:
        row_key = "|".join(str(row[key]) for key in (
            "auction_id", "network_point_id", "direction", "flow_start", "flow_end"
        ))
        previous_exit = row["exit_market"]
        previous_entry = row["entry_market"]

        def result(status, code, message, proposed_exit=previous_exit,
                   proposed_entry=previous_entry, final_exit=previous_exit,
                   final_entry=previous_entry, changed=False):
            return HistoricalBackfillAudit(
                run_id, row["id"], row_position, str(row["auction_id"]), row_key,
                previous_exit, previous_entry, proposed_exit, proposed_entry,
                final_exit, final_entry, status, code, message, changed,
            )

        if not AuctionStorage._valid_historical_row(row):
            return result(HistoricalBackfillStatus.INVALID, "invalid_historical_row",
                          "The stored row has invalid backfill coordinates.")
        if row["direction"] == "bundle":
            return result(HistoricalBackfillStatus.UNRESOLVABLE,
                          "insufficient_bundle_identity",
                          "The stored bundle row does not retain both original side identities.")
        sides = {
            "exit": (ReferenceSide.EXIT,),
            "entry": (ReferenceSide.ENTRY,),
        }[row["direction"]]
        references = {side: catalog.lookup(row["network_point"], side) for side in sides}
        if any(reference is None for reference in references.values()):
            return result(HistoricalBackfillStatus.UNRESOLVABLE, "reference_unresolvable",
                          "The stored network point cannot be resolved for every required side.")
        proposed_exit = references[ReferenceSide.EXIT].canonical_name if ReferenceSide.EXIT in sides else previous_exit
        proposed_entry = references[ReferenceSide.ENTRY].canonical_name if ReferenceSide.ENTRY in sides else previous_entry
        missing = lambda value: value is None or (isinstance(value, str) and not value.strip())
        equivalent = lambda value, canonical: (
            isinstance(value, str) and value.strip().casefold() == canonical.casefold()
        )
        conflicts = (
            (ReferenceSide.EXIT in sides and not missing(previous_exit)
             and not equivalent(previous_exit, proposed_exit))
            or (ReferenceSide.ENTRY in sides and not missing(previous_entry)
                and not equivalent(previous_entry, proposed_entry))
        )
        if conflicts:
            return result(HistoricalBackfillStatus.CONFLICT, "reference_conflict",
                          "A stored non-empty value conflicts with the reference-derived value.",
                          proposed_exit, proposed_entry)
        final_exit = proposed_exit if ReferenceSide.EXIT in sides and missing(previous_exit) else previous_exit
        final_entry = proposed_entry if ReferenceSide.ENTRY in sides and missing(previous_entry) else previous_entry
        if (final_exit, final_entry) == (previous_exit, previous_entry):
            return result(HistoricalBackfillStatus.ALREADY_COMPLETE, "already_complete",
                          "All required Market / Storage values are already complete.")
        return result(HistoricalBackfillStatus.UPDATED, "missing_values_filled",
                      "Missing Market / Storage values were filled from the reference catalog.",
                      proposed_exit, proposed_entry, final_exit, final_entry, True)

    @staticmethod
    def _valid_historical_row(row: sqlite3.Row) -> bool:
        text_fields = ("auction_id", "network_point", "network_point_id", "auction_date",
                       "flow_start", "flow_end", "product_type")
        if row["direction"] not in {"exit", "entry", "bundle"} or any(
            not isinstance(row[key], str) or not row[key].strip() for key in text_fields
        ):
            return False
        try:
            auction_date = datetime.fromisoformat(row["auction_date"])
            flow_start = datetime.fromisoformat(row["flow_start"])
            flow_end = datetime.fromisoformat(row["flow_end"])
        except (TypeError, ValueError):
            return False
        if (flow_start.utcoffset() is None) != (flow_end.utcoffset() is None):
            return False
        try:
            invalid_order = flow_start >= flow_end
        except (TypeError, ValueError):
            return False
        if invalid_order or not isinstance(auction_date, datetime):
            return False
        for key in ("booked_capacity_kwh_h", "runtime_hours", "tariff_eur_mwh_h",
                    "premium_eur_mwh_h"):
            value = row[key]
            if isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(value):
                return False
        return True

    @staticmethod
    def _validate_historical_backfill(
        connection: sqlite3.Connection, audit: list[HistoricalBackfillAudit], run_id: str
    ) -> None:
        if len({item.auction_row_id for item in audit}) != len(audit):
            raise AuctionStorageError("Historical backfill audit row identifiers are not unique.")
        if len(audit) != sum(
            item.status is status
            for status in HistoricalBackfillStatus
            for item in audit
        ):
            raise AuctionStorageError("Historical backfill summary accounting failed.")
        changed_ids = [item.auction_row_id for item in audit if item.changed]
        for item in audit:
            stored = connection.execute(
                "SELECT exit_market, entry_market FROM auctions WHERE id=?",
                (item.auction_row_id,),
            ).fetchone()
            if stored is None or (stored["exit_market"], stored["entry_market"]) != (
                item.final_exit_market,
                item.final_entry_market,
            ):
                raise AuctionStorageError("Historical backfill validation failed.")
        if len(changed_ids) != sum(item.changed for item in audit):
            raise AuctionStorageError("Historical backfill change accounting failed.")
        persisted = connection.execute(
            "SELECT examined, updated, unchanged, skipped, conflicts, invalid "
            "FROM historical_market_storage_runs WHERE run_id=?", (run_id,)
        ).fetchone()
        if persisted is None or tuple(persisted) != (len(audit), *AuctionStorage._historical_backfill_counts(audit)):
            raise AuctionStorageError("Historical backfill run summary validation failed.")
        if connection.execute(
            "SELECT count(*) FROM historical_market_storage_audit WHERE run_id=?", (run_id,)
        ).fetchone()[0] != len(audit):
            raise AuctionStorageError("Historical backfill persisted audit validation failed.")

    def operations(self) -> list[sqlite3.Row]:
        with self._connection() as connection, connection:
            return list(connection.execute(
                "SELECT * FROM prisma_source_operations ORDER BY source_date"
            ))

    def unresolved_operations(self) -> list[sqlite3.Row]:
        return [row for row in self.operations() if row["status"] != "accepted"]

    def import_legacy_operation(
        self, operation_id: str, source_date: str, source_name: str, digest: str
    ) -> None:
        with self._connection() as connection, connection:
            connection.execute(
                "INSERT OR IGNORE INTO prisma_source_operations "
                "(operation_id, source_date, source_name, sha256, status) "
                "VALUES (?, ?, ?, ?, 'accepted')",
                (operation_id, source_date, source_name, digest),
            )

    def operation_for_date(self, source_date: str) -> sqlite3.Row | None:
        with self._connection() as connection, connection:
            return connection.execute(
                "SELECT * FROM prisma_source_operations WHERE source_date = ?", (source_date,)
            ).fetchone()

    def begin_operation(self, source_date: str, source_name: str, digest: str) -> sqlite3.Row:
        existing = self.operation_for_date(source_date)
        if existing is not None:
            if existing["sha256"] != digest:
                state = "accepted" if existing["status"] == "accepted" else "unresolved"
                raise AuctionStorageError(
                    f"A different PRISMA source for this date is already {state}."
                )
            return existing
        operation_id = uuid.uuid4().hex
        with self._connection() as connection, connection:
            connection.execute(
                "INSERT INTO prisma_source_operations "
                "(operation_id, source_date, source_name, sha256, status) VALUES (?, ?, ?, ?, 'pending')",
                (operation_id, source_date, source_name, digest),
            )
        return self.operation_for_date(source_date)  # type: ignore[return-value]

    def apply_operation(self, operation_id: str, rows: list[dict[str, Any]], summary: dict[str, Any]) -> dict[str, int]:
        with self._connection() as connection, connection:
            operation = connection.execute(
                "SELECT status, summary_json FROM prisma_source_operations WHERE operation_id = ?",
                (operation_id,),
            ).fetchone()
            if operation is None:
                raise AuctionStorageError("The pending PRISMA operation was not found.")
            if operation["status"] != "pending":
                stored = json.loads(operation["summary_json"] or "{}")
                return {key: int(stored[key]) for key in ("processed", "inserted", "updated", "unchanged")}
            stats = self._upsert_rows(connection, rows)
            summary.update(stats)
            connection.execute(
                "UPDATE prisma_source_operations SET status='data_committed', summary_json=?, "
                "updated_at=CURRENT_TIMESTAMP WHERE operation_id=? AND status='pending'",
                (json.dumps(summary, sort_keys=True), operation_id),
            )
        return stats

    @staticmethod
    def _validate_upsert_batch(rows: list[dict[str, Any]]) -> None:
        rows_by_identity: dict[tuple[Any, ...], dict[str, Any]] = {}
        for row in rows:
            network_point_id = row.get("network_point_id")
            if not isinstance(network_point_id, str) or not network_point_id.strip():
                raise AuctionStorageError(
                    "Auction network_point_id must be a nonblank string."
                )
            identity = tuple(row.get(field) for field in AuctionStorage.AUCTION_IDENTITY_FIELDS)
            previous = rows_by_identity.get(identity)
            if previous is not None and any(
                previous.get(field) != row.get(field)
                for field in AuctionStorage.AUCTION_PERSISTED_FIELDS
            ):
                raise AuctionStorageError(
                    "The auction batch contains conflicting rows with the same identity."
                )
            rows_by_identity[identity] = row

    @staticmethod
    def _upsert_rows(
        connection: sqlite3.Connection, rows: list[dict[str, Any]]
    ) -> dict[str, int]:
        AuctionStorage._validate_upsert_batch(rows)
        inserted = updated = unchanged = 0
        for row in rows:
            existing = connection.execute(
                "SELECT * FROM auctions WHERE auction_id=? AND network_point_id=? "
                "AND direction=? AND flow_start=? AND flow_end=?",
                (row["auction_id"], row["network_point_id"], row["direction"],
                 row["flow_start"], row["flow_end"]),
            ).fetchone()
            if existing is None:
                columns = ", ".join(row)
                connection.execute(
                    f"INSERT INTO auctions ({columns}) VALUES ({', '.join('?' for _ in row)})",
                    tuple(row.values()),
                )
                inserted += 1
            elif any(
                existing[key] != value for key, value in row.items()
                if key in existing.keys()
            ):
                assignments = ", ".join(f"{key}=?" for key in row)
                connection.execute(
                    f"UPDATE auctions SET {assignments}, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                    (*row.values(), existing["id"]),
                )
                updated += 1
            else:
                unchanged += 1
        return {"processed": len(rows), "inserted": inserted,
                "updated": updated, "unchanged": unchanged}

    def finalize_operation(self, operation_id: str) -> None:
        with self._connection() as connection, connection:
            changed = connection.execute(
                "UPDATE prisma_source_operations SET status='accepted', updated_at=CURRENT_TIMESTAMP "
                "WHERE operation_id=? AND status='data_committed'", (operation_id,)
            ).rowcount
            if changed != 1:
                raise AuctionStorageError("The PRISMA operation could not be finalized safely.")

    @staticmethod
    def apply_excel_widths(path: Path) -> None:
        workbook = load_workbook(path)
        try:
            sheet = workbook["Auctions"]
            for index, header in enumerate(AuctionStorage.EXCEL_COLUMNS, start=1):
                sheet.column_dimensions[get_column_letter(index)].width = (
                    AuctionStorage.EXCEL_COLUMN_WIDTHS[header]
                )
            workbook.save(path)
        finally:
            workbook.close()

    @staticmethod
    def validate_excel(path: Path) -> bool:
        workbook = None
        try:
            workbook = load_workbook(path, read_only=False, data_only=True)
            valid = False
            if "Auctions" in workbook.sheetnames:
                sheet = workbook["Auctions"]
                headers = tuple(
                    cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))
                )
                widths_are_valid = all(
                    abs(
                        sheet.column_dimensions[get_column_letter(index)].width
                        - AuctionStorage.EXCEL_COLUMN_WIDTHS[header]
                    ) <= AuctionStorage.EXCEL_WIDTH_TOLERANCE
                    for index, header in enumerate(AuctionStorage.EXCEL_COLUMNS, start=1)
                )
                valid = headers == AuctionStorage.EXCEL_COLUMNS and widths_are_valid
            return valid
        except Exception:
            return False
        finally:
            if workbook is not None:
                workbook.close()

    def export_excel(self, output_path: Path) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with self._connection() as connection, connection:
            frame = pd.read_sql_query("""
                SELECT auction_date AS "Auction Date", exit_market AS "Exit Market/Storage",
                entry_market AS "Entry Market/Storage", direction AS "Capacity Type",
                network_point AS "Network Point Name", product_type AS "Product Type",
                flow_start AS "Flow Start", flow_end AS "Flow End",
                booked_capacity_kwh_h AS "Booked Capacity, kWh/h", runtime_hours AS "Runtime Hours",
                tariff_eur_mwh_h AS "Tariff, EUR/MWh/h", premium_eur_mwh_h AS "Premium, EUR/MWh/h",
                auction_id AS "Auction ID", tso_exit AS "TSO Exit", tso_entry AS "TSO Entry", state AS "Status"
                FROM auctions ORDER BY auction_date, auction_id, network_point_id, direction, flow_start, flow_end
            """, connection)
        staged: Path | None = None
        try:
            descriptor, name = tempfile.mkstemp(prefix=f".{output_path.stem}-", suffix=".xlsx", dir=output_path.parent)
            os.close(descriptor)
            staged = Path(name)
            frame.to_excel(staged, index=False, sheet_name="Auctions")
            self.apply_excel_widths(staged)
            if not self.validate_excel(staged):
                raise AuctionStorageError("The staged Excel workbook failed validation.")
            try:
                os.replace(staged, output_path)
            except PermissionError as exc:
                raise AuctionStorageError(
                    "The Excel output is open or locked. Close it and retry the import."
                ) from exc
            staged = None
        except AuctionStorageError:
            raise
        except Exception as exc:
            raise AuctionStorageError("The Excel output could not be staged safely.") from exc
        finally:
            if staged is not None:
                try:
                    staged.unlink(missing_ok=True)
                except OSError:
                    pass
        return output_path

    def upsert(self, rows: list[dict[str, Any]]) -> dict[str, int]:
        """Compatibility API for storage-only callers."""
        with self._connection() as connection, connection:
            return self._upsert_rows(connection, rows)
